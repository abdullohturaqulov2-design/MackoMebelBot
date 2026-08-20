# -*- coding: utf-8 -*-
"""
Excel orqali kategoriya/subkategoriya qo'shish, o'chirish, tahrirlash.
Shablon formati:
  Tur | Slug | Nomi (uz) | Nomi (ru) | Nomi (en) | Ota slug | Amal
  kategoriya | akril | Akril | Акрил | Acrylic | | qoshish
  subkategoriya | stalish | Stalishnitsa | Столешница | Countertop | akril | qoshish
  subkategoriya | old_slug | | | | | ochirish
"""
import os, uuid
from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from aiogram.fsm.context import FSMContext
from database import db
from keyboards.reply import get_main_keyboard, cancel_only_keyboard
from locales.texts import TEXTS, t
from states.states import AdminStates
from utils.access import is_admin, get_role_level
from utils.render import delete_msg
from config import DATA_DIR

router = Router()
def _all(k): return {TEXTS[l].get(k,"") for l in TEXTS} - {""}


@router.message(F.text.in_(_all("btn_manage_cats")))
async def cat_menu_excel_option(message: Message, state: FSMContext):
    """Kategoriya menyusida Excel tugmasini ko'rsatish — warehouse handler bilan ishlamaydi."""
    pass  # Bu categories.py da boshqariladi


@router.callback_query(F.data == "cmgr:excel")
async def cb_cat_excel(callback, state: FSMContext):
    lang = db.get_user_lang(callback.from_user.id)
    if not is_admin(callback.from_user.id):
        await callback.answer(); return
    await state.set_state(AdminStates.cat_excel)
    await callback.message.answer(
        "📊 <b>Kategoriya Excel boshqaruvi</b>\n\n"
        "Excel faylni yuboring.\n\n"
        "<b>Format (ustunlar tartibi):</b>\n"
        "1. Tur: <code>kategoriya</code> yoki <code>subkategoriya</code>\n"
        "2. Slug: <code>akril</code> (lotin, _)\n"
        "3. Nomi (uz): <code>Akril</code>\n"
        "4. Nomi (ru): <code>Акрил</code> (bo'sh = uz nom)\n"
        "5. Nomi (en): <code>Acrylic</code> (bo'sh = uz nom)\n"
        "6. Ota slug: <code>akril</code> (subkategoriya uchun)\n"
        "7. Amal: <code>qoshish</code> yoki <code>ochirish</code>\n\n"
        "📥 Shablon faylni yuklab oling va to'ldiring:",
        parse_mode="HTML",
        reply_markup=cancel_only_keyboard(lang))
    # Shablon yaratish va yuborish
    await _send_cat_template(callback.message)
    await callback.answer()


async def _send_cat_template(message):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Kategoriyalar"
    DARK = PatternFill("solid", fgColor="0D2137")
    BLUE = PatternFill("solid", fgColor="2E75B6")
    SMPL = PatternFill("solid", fgColor="DEEAF1")

    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = "📂 KATEGORIYA BOSHQARUVI SHABLONI"
    c.font = Font(bold=True, size=12, color="FFFFFF"); c.fill = DARK
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    headers = ["Tur*","Slug*","Nomi (uz)*","Nomi (ru)","Nomi (en)","Ota slug","Amal*"]
    widths  = [16, 18, 22, 22, 22, 18, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = BLUE; cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64+i)].width = w
    ws.row_dimensions[2].height = 22

    samples = [
        ("kategoriya","akril","Akril","Акрил","Acrylic","","qoshish"),
        ("kategoriya","mdf","MDF","МДФ","MDF","","qoshish"),
        ("subkategoriya","stalish","Stalishnitsa","Столешница","Countertop","akril","qoshish"),
        ("subkategoriya","premium","Premium akril","Премиум акрил","Premium acrylic","stalish","qoshish"),
        ("kategoriya","eski","Eski kategoriya","","","","ochirish"),
    ]
    for r, row in enumerate(samples, 3):
        for c2, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c2, value=val)
            cell.fill = SMPL

    for i in range(len(samples)+3, len(samples)+33):
        for j in range(1, 8): ws.cell(row=i, column=j, value="")

    path = os.path.join(DATA_DIR, "tmp", f"cat_tmpl_{uuid.uuid4().hex[:6]}.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    await message.answer_document(
        FSInputFile(path, filename="shablon_kategoriya.xlsx"),
        caption="📂 Mana shablon. To'ldiring va yuboring.")
    try: os.remove(path)
    except: pass


@router.message(AdminStates.cat_excel, F.document)
async def process_cat_excel(message: Message, state: FSMContext, bot):
    if not is_admin(message.from_user.id): return
    lang = db.get_user_lang(message.from_user.id)
    if not (message.document.file_name or "").lower().endswith((".xlsx",".xlsm")):
        await message.answer("⚠️ Faqat Excel (.xlsx) fayl!"); return

    wait = await message.answer("⏳ Ishlanmoqda...")
    tmp  = os.path.join(DATA_DIR, "tmp"); os.makedirs(tmp, exist_ok=True)
    path = os.path.join(tmp, f"cat_{uuid.uuid4().hex}.xlsx")
    fi   = await bot.get_file(message.document.file_id)
    await bot.download_file(fi.file_path, destination=path)

    import openpyxl
    from utils.excel_utils import _nh
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Header qatorni topish
    header_row = 0
    for i, row in enumerate(rows[:5]):
        vals = [str(v or "").lower() for v in row]
        if any("tur" in v or "slug" in v for v in vals):
            header_row = i; break

    added = updated = deleted = errors = 0
    result_lines = []

    for row in rows[header_row+1:]:
        if not row or not row[0]: continue
        typ   = str(row[0] or "").strip().lower()
        slug  = str(row[1] or "").strip().lower()
        uz    = str(row[2] or "").strip()
        ru    = str(row[3] or "").strip() or uz
        en    = str(row[4] or "").strip() or uz
        parent= str(row[5] or "").strip().lower()
        amal  = str(row[6] or "").strip().lower()

        if not slug or not typ: continue

        try:
            if amal in ("ochirish","o'chirish","delete","del"):
                if typ == "kategoriya":
                    ok = db.delete_category(slug)
                    if ok: deleted += 1; result_lines.append(f"🗑 Kategoriya o'chirildi: {slug}")
                    else: result_lines.append(f"⚠️ O'chirib bo'lmadi (mahsulot bor): {slug}")
                else:
                    ok = db.delete_subcategory(parent, slug)
                    if ok: deleted += 1; result_lines.append(f"🗑 Subkategoriya o'chirildi: {slug}")
            elif typ == "kategoriya":
                st = db.upsert_category_from_excel(slug, uz, ru, en)
                if st == "added": added += 1; result_lines.append(f"✅ Yangi: {uz}")
                else: updated += 1; result_lines.append(f"🔄 Yangilandi: {uz}")
            elif "sub" in typ:
                if parent:
                    # Avval parent subcategory id topish (nested uchun)
                    parent_sub = db.get_subcategory(slug.split("_")[0] if "_" in parent else "", parent)
                    if parent_sub:
                        db.add_nested_sub(slug.split("_")[0] if "_" in parent else parent,
                                          parent_sub["id"], slug, uz)
                        added += 1; result_lines.append(f"✅ Sub-sub: {uz}")
                    else:
                        # Oddiy subkategoriya
                        st = db.upsert_subcategory_from_excel(parent, slug, uz, ru, en)
                        if st == "added": added += 1; result_lines.append(f"✅ Sub yangi: {uz}")
                        else: updated += 1; result_lines.append(f"🔄 Sub yangilandi: {uz}")
                else:
                    errors += 1; result_lines.append(f"⚠️ Ota slug yo'q: {slug}")
        except Exception as e:
            errors += 1; result_lines.append(f"❌ Xato ({slug}): {e}")

    try: os.remove(path)
    except: pass
    await wait.delete()

    summary = (f"📊 <b>Natija:</b>\n"
               f"✅ Yangi: {added} ta\n"
               f"🔄 Yangilandi: {updated} ta\n"
               f"🗑 O'chirildi: {deleted} ta\n"
               f"❌ Xato: {errors} ta\n\n")
    if result_lines:
        summary += "\n".join(result_lines[:15])
        if len(result_lines) > 15:
            summary += f"\n... va yana {len(result_lines)-15} ta"

    role = get_role_level(message.from_user.id)
    await state.set_state(AdminStates.main_menu)
    await message.answer(summary, parse_mode="HTML")
    await message.answer(t(lang,"main_menu"), reply_markup=get_main_keyboard(lang, role))

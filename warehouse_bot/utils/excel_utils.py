# -*- coding: utf-8 -*-
import re, os
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl import Workbook
from config import IMAGES_DIR

HEADER_MAP = {
    "nomi":"name","name":"name","название":"name","наименование":"name",
    "kodi":"code","kod":"code","code":"code","код":"code",
    "kategoriya":"category","category":"category","категория":"category",
    "subkategoriya":"subcategory","subcategory":"subcategory","подкатегория":"subcategory",
    "format":"format_size","формат":"format_size",
    "qalinlik":"thickness","qalinligi":"thickness","thickness":"thickness","толщина":"thickness",
    "narx":"price","narxi":"price","price":"price","цена":"price","cost":"price",
    "miqdor":"quantity","miqdori":"quantity","quantity":"quantity","количество":"quantity",
    "joylashuv":"location","joy":"location","location":"location","место":"location",
    "rasm":"image_path","rasm yoli":"image_path","image":"image_path","изображение":"image_path",
}


def _nh(v) -> str:
    """
    Header nomini normallashtiradi.
    "Nomi ★" → "nomi"
    "Narx ★★ (so'm)" → "narx"
    "Kategoriya ★" → "kategoriya"
    """
    if not v: return ""
    s = str(v).strip().lower()
    s = re.sub(r"\(.*?\)", "", s)        # (so'm) kabi qavslarni olib tashlash
    s = re.sub(r"[^\w\s]", " ", s)      # ★ ★★ va boshqa belgilarni bo'sh joy bilan
    s = re.sub(r"\s+", " ", s).strip()  # ko'p bo'sh joyni bittaga
    return s


def _resolve_image(raw: str):
    if not raw: return None
    raw = raw.strip()
    if os.path.isabs(raw) or os.path.sep in raw or "/" in raw:
        return raw if os.path.exists(raw) else None
    candidate = os.path.join(IMAGES_DIR, raw)
    if os.path.exists(candidate): return candidate
    for ext in (".jpg",".jpeg",".png",".webp",".JPG",".PNG"):
        c = os.path.join(IMAGES_DIR, raw + ext)
        if os.path.exists(c): return c
    return None


def _find_header_row(rows) -> int:
    """
    Sarlavha qatorini topadi (title qatorlarini o'tkazib yuboradi).
    Misol: "OMBOR MAHSULOTLARINI QO'SHISH SHABLONI" — title, o'tkaziladi.
    "Nomi, Kodi, Kategoriya..." — sarlavha qatori.
    """
    for i, row in enumerate(rows[:15]):
        normalized = {_nh(cell) for cell in row if cell}
        matched    = normalized.intersection(set(HEADER_MAP.keys()))
        if len(matched) >= 2:   # kamida 2 ta taniqli ustun
            return i
    return 0


def read_products_from_excel(path: str) -> Tuple[List[Dict], List[str]]:
    from database.db import category_exists
    wb   = openpyxl.load_workbook(path, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], ["Fayl bo'sh."]

    # Sarlavha qatorini avtomatik topish (title/izoh qatorlarini o'tkazadi)
    header_idx = _find_header_row(rows)
    header_row = rows[header_idx]
    data_rows  = rows[header_idx + 1:]

    col_map = {}
    for i, cell in enumerate(header_row):
        k = _nh(cell)
        if k in HEADER_MAP:
            col_map[HEADER_MAP[k]] = i

    if "name" not in col_map or "category" not in col_map:
        found = [_nh(c) for c in header_row if c]
        return [], [
            f"'Nomi' va 'Kategoriya' ustunlari topilmadi.\n"
            f"Topilgan ustunlar: {found}\n"
            f"Sarlavha qatori: {header_idx + 1}-qator"
        ]

    products, errors = [], []
    for i, row in enumerate(data_rows, header_idx + 2):
        def g(f):
            idx = col_map.get(f)
            if idx is None or idx >= len(row) or row[idx] is None: return ""
            return str(row[idx]).strip()

        name     = g("name")
        category = g("category").lower()
        if not name:
            errors.append(f"{i}-qator: nomi bo'sh, o'tkazildi."); continue
        if not category:
            errors.append(f"{i}-qator ({name}): kategoriya bo'sh."); continue
        if not category_exists(category):
            errors.append(f"{i}-qator ({name}): '{category}' kategoriyasi botda yo'q."); continue

        def to_float(val):
            try: return float(val.replace(" ","").replace(",",".")) if val else 0.0
            except: return 0.0

        products.append({
            "name": name, "code": g("code"),
            "category": category, "subcategory": g("subcategory"),
            "format_size": g("format_size"), "thickness": g("thickness"),
            "quantity": to_float(g("quantity")),
            "price":    to_float(g("price")),
            "location": g("location"),
            "image_path": _resolve_image(g("image_path")) if g("image_path") else None,
        })
    return products, errors


def read_delete_codes_from_excel(path: str) -> List[str]:
    wb = openpyxl.load_workbook(path, data_only=True)

    # OCHIRISH sheetini qidirish (Delete/Ochirish/ochir nomli sheet)
    ws = None
    delete_keywords = ["ochir", "delete", "del", "удал", "remove"]
    for name in wb.sheetnames:
        if any(k in name.lower() for k in delete_keywords):
            ws = wb[name]; break
    if ws is None:
        ws = wb.active  # topilmasa birinchi sheet

    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []

    # Sarlavha qatorini topish (title qatorlarini o'tkazib yuborish)
    header_idx = _find_header_row(rows)
    header_row = rows[header_idx]
    data_rows  = rows[header_idx + 1:]

    # Kodi ustunini topish
    code_idx = 0
    for i, cell in enumerate(header_row):
        if _nh(cell) in ("kodi","kod","code","код"):
            code_idx = i; break

    codes = []
    for row in data_rows:
        if code_idx < len(row) and row[code_idx]:
            val = str(row[code_idx]).strip()
            if val: codes.append(val)
    return codes


def export_products_to_excel(products: List[Dict], path: str, title: str = "Mahsulotlar"):
    wb = Workbook(); ws = wb.active; ws.title = title[:31]
    ws.append(["Nomi","Kodi","Kategoriya","Subkategoriya",
                "Format","Qalinlik","Narx (so'm)","Miqdor","Joylashuv","Qo'shilgan"])
    for p in products:
        ws.append([
            p.get("name",""), p.get("code",""), p.get("category",""),
            p.get("subcategory",""), p.get("format_size",""), p.get("thickness",""),
            p.get("price",0), p.get("quantity",""), p.get("location",""), p.get("added_at",""),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col)+2, 40)
    wb.save(path)
